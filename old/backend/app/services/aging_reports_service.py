"""
Enhanced Aging Reports service
"""
from bson import ObjectId
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
import logging
from io import BytesIO
import json

from app.database import Database
from app.api.exceptions import NotFoundException, BadRequestException

logger = logging.getLogger(__name__)


class AgingReportsService:
    """Service for enhanced aging reports"""
    
    @staticmethod
    def get_detailed_ar_aging(
        tenant_id: str,
        as_of_date: Optional[str] = None,
        client_id: Optional[str] = None,
        include_zero_balances: bool = False,
        min_amount: Optional[float] = None,
        max_amount: Optional[float] = None
    ) -> Dict:
        """
        Get detailed accounts receivable aging report
        
        Returns:
            Dict with detailed aging data
        """
        db = Database.get_db()
        
        if not as_of_date:
            as_of_date = datetime.utcnow().strftime("%Y-%m-%d")
        
        as_of_datetime = datetime.strptime(as_of_date, "%Y-%m-%d")
        
        # Build query
        query = {
            "tenant_id": tenant_id,
            "status": {"$in": ["sent", "overdue"]},
            "amount_due": {"$gt": 0}
        }
        
        if client_id:
            query["client_id"] = client_id
        
        # Get all outstanding invoices
        invoices = list(db.invoices.find(query))
        
        # Calculate aging buckets
        aging_data = {
            "as_of_date": as_of_date,
            "total_outstanding": 0.0,
            "buckets": {
                "current": {"label": "Current (0-30 days)", "amount": 0.0, "count": 0, "invoices": []},
                "30_days": {"label": "31-60 days", "amount": 0.0, "count": 0, "invoices": []},
                "60_days": {"label": "61-90 days", "amount": 0.0, "count": 0, "invoices": []},
                "90_plus": {"label": "90+ days", "amount": 0.0, "count": 0, "invoices": []}
            },
            "clients": {},
            "summary": {
                "total_invoices": 0,
                "total_clients": 0,
                "average_days_outstanding": 0.0,
                "largest_outstanding": 0.0,
                "oldest_invoice_days": 0
            }
        }
        
        client_totals = {}
        total_days_outstanding = 0
        largest_outstanding = 0.0
        oldest_days = 0
        
        for invoice in invoices:
            # Get client information
            client = db.clients.find_one({"_id": ObjectId(invoice["client_id"])})
            client_name = client.get("name", "Unknown Client") if client else "Unknown Client"
            
            # Calculate days outstanding
            due_date = datetime.strptime(invoice["due_date"], "%Y-%m-%d")
            days_outstanding = (as_of_datetime - due_date).days
            
            amount_due = invoice.get("amount_due", 0.0)
            
            # Apply filters
            if min_amount and amount_due < min_amount:
                continue
            if max_amount and amount_due > max_amount:
                continue
            if not include_zero_balances and amount_due <= 0:
                continue
            
            # Determine aging bucket
            if days_outstanding <= 30:
                bucket = "current"
            elif days_outstanding <= 60:
                bucket = "30_days"
            elif days_outstanding <= 90:
                bucket = "60_days"
            else:
                bucket = "90_plus"
            
            # Create invoice summary
            invoice_summary = {
                "id": str(invoice["_id"]),
                "invoice_number": invoice.get("invoice_number"),
                "client_id": invoice["client_id"],
                "client_name": client_name,
                "invoice_date": invoice.get("invoice_date"),
                "due_date": invoice["due_date"],
                "total": invoice.get("total", 0.0),
                "amount_paid": invoice.get("amount_paid", 0.0),
                "amount_due": amount_due,
                "days_outstanding": days_outstanding,
                "status": invoice.get("status"),
                "notes": invoice.get("notes", "")
            }
            
            # Add to bucket
            aging_data["buckets"][bucket]["amount"] += amount_due
            aging_data["buckets"][bucket]["count"] += 1
            aging_data["buckets"][bucket]["invoices"].append(invoice_summary)
            
            # Add to client totals
            if client_name not in client_totals:
                client_totals[client_name] = {
                    "client_id": invoice["client_id"],
                    "client_name": client_name,
                    "total_outstanding": 0.0,
                    "invoice_count": 0,
                    "average_days_outstanding": 0.0,
                    "invoices": [],
                    "buckets": {
                        "current": 0.0,
                        "30_days": 0.0,
                        "60_days": 0.0,
                        "90_plus": 0.0
                    }
                }
            
            client_totals[client_name]["total_outstanding"] += amount_due
            client_totals[client_name]["invoice_count"] += 1
            client_totals[client_name]["invoices"].append(invoice_summary)
            client_totals[client_name]["buckets"][bucket] += amount_due
            
            # Update summary statistics
            aging_data["total_outstanding"] += amount_due
            total_days_outstanding += days_outstanding
            
            if amount_due > largest_outstanding:
                largest_outstanding = amount_due
            
            if days_outstanding > oldest_days:
                oldest_days = days_outstanding
        
        # Calculate client averages
        for client_data in client_totals.values():
            if client_data["invoice_count"] > 0:
                total_client_days = sum(inv["days_outstanding"] for inv in client_data["invoices"])
                client_data["average_days_outstanding"] = total_client_days / client_data["invoice_count"]
        
        aging_data["clients"] = client_totals
        
        # Calculate summary statistics
        total_invoices = sum(bucket["count"] for bucket in aging_data["buckets"].values())
        aging_data["summary"]["total_invoices"] = total_invoices
        aging_data["summary"]["total_clients"] = len(client_totals)
        aging_data["summary"]["largest_outstanding"] = largest_outstanding
        aging_data["summary"]["oldest_invoice_days"] = oldest_days
        
        if total_invoices > 0:
            aging_data["summary"]["average_days_outstanding"] = total_days_outstanding / total_invoices
        
        return aging_data
    
    @staticmethod
    def get_client_aging_details(
        tenant_id: str,
        client_id: str,
        as_of_date: Optional[str] = None
    ) -> Dict:
        """
        Get detailed aging information for a specific client
        
        Returns:
            Dict with client aging details
        """
        db = Database.get_db()
        
        # Get client information
        client = db.clients.find_one({"_id": ObjectId(client_id)})
        if not client:
            raise NotFoundException("Client not found")
        
        # Get aging data for this client
        aging_data = AgingReportsService.get_detailed_ar_aging(
            tenant_id=tenant_id,
            client_id=client_id,
            as_of_date=as_of_date,
            include_zero_balances=False
        )
        
        client_name = client.get("name", "Unknown Client")
        client_data = aging_data["clients"].get(client_name, {})
        
        # Get payment history
        payment_history = AgingReportsService._get_client_payment_history(
            db, tenant_id, client_id, limit=10
        )
        
        # Get collection notes
        collection_notes = AgingReportsService._get_collection_notes(
            db, tenant_id, client_id
        )
        
        return {
            "client": {
                "id": str(client["_id"]),
                "name": client_name,
                "email": client.get("email"),
                "phone": client.get("phone"),
                "address": client.get("address")
            },
            "aging_summary": client_data,
            "payment_history": payment_history,
            "collection_notes": collection_notes,
            "as_of_date": aging_data["as_of_date"]
        }
    
    @staticmethod
    def add_collection_note(
        tenant_id: str,
        client_id: Optional[str] = None,
        invoice_id: Optional[str] = None,
        note_text: str = "",
        note_type: str = "general",
        created_by: str = "system"
    ) -> Dict:
        """
        Add a collection note for a client or invoice
        
        Returns:
            Dict with created note data
        """
        db = Database.get_db()
        
        if not client_id and not invoice_id:
            raise BadRequestException("Either client_id or invoice_id must be provided")
        
        # If invoice_id provided, get client_id from invoice
        if invoice_id and not client_id:
            invoice = db.invoices.find_one({"_id": ObjectId(invoice_id)})
            if invoice:
                client_id = invoice.get("client_id")
        
        # Create collection note
        note = {
            "tenant_id": tenant_id,
            "client_id": client_id,
            "invoice_id": invoice_id,
            "note_text": note_text,
            "note_type": note_type,  # general, phone_call, email, payment_plan, legal_action
            "created_by": created_by,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        
        result = db.collection_notes.insert_one(note)
        
        # Return note with string ID
        note["id"] = str(result.inserted_id)
        del note["_id"]
        
        return note
    
    @staticmethod
    def get_collection_notes(
        tenant_id: str,
        client_id: Optional[str] = None,
        invoice_id: Optional[str] = None,
        limit: int = 50
    ) -> List[Dict]:
        """
        Get collection notes for a client or invoice
        
        Returns:
            List of collection notes
        """
        db = Database.get_db()
        
        query = {"tenant_id": tenant_id}
        
        if client_id:
            query["client_id"] = client_id
        
        if invoice_id:
            query["invoice_id"] = invoice_id
        
        notes = list(db.collection_notes.find(query)
                    .sort("created_at", -1)
                    .limit(limit))
        
        # Convert ObjectId to string
        for note in notes:
            note["id"] = str(note["_id"])
            del note["_id"]
        
        return notes
    
    @staticmethod
    def update_collection_note(
        tenant_id: str,
        note_id: str,
        note_text: Optional[str] = None,
        note_type: Optional[str] = None
    ) -> Dict:
        """
        Update a collection note
        
        Returns:
            Dict with updated note data
        """
        db = Database.get_db()
        
        update_data = {"updated_at": datetime.utcnow()}
        
        if note_text is not None:
            update_data["note_text"] = note_text
        
        if note_type is not None:
            update_data["note_type"] = note_type
        
        result = db.collection_notes.update_one(
            {
                "_id": ObjectId(note_id),
                "tenant_id": tenant_id
            },
            {"$set": update_data}
        )
        
        if result.modified_count == 0:
            raise NotFoundException("Collection note not found")
        
        # Return updated note
        note = db.collection_notes.find_one({"_id": ObjectId(note_id)})
        note["id"] = str(note["_id"])
        del note["_id"]
        
        return note
    
    @staticmethod
    def delete_collection_note(
        tenant_id: str,
        note_id: str
    ) -> Dict:
        """
        Delete a collection note
        
        Returns:
            Dict with success message
        """
        db = Database.get_db()
        
        result = db.collection_notes.delete_one({
            "_id": ObjectId(note_id),
            "tenant_id": tenant_id
        })
        
        if result.deleted_count == 0:
            raise NotFoundException("Collection note not found")
        
        return {"message": "Collection note deleted successfully"}
    
    @staticmethod
    def setup_automated_reminders(
        tenant_id: str,
        reminder_schedule: List[Dict],
        email_template: Optional[str] = None,
        sms_template: Optional[str] = None,
        enabled: bool = True
    ) -> Dict:
        """
        Setup automated reminder configuration
        
        Args:
            reminder_schedule: List of reminder configs like [{"days_overdue": 7, "method": "email"}, ...]
        
        Returns:
            Dict with reminder configuration
        """
        db = Database.get_db()
        
        # Create or update reminder configuration
        config = {
            "tenant_id": tenant_id,
            "reminder_schedule": reminder_schedule,
            "email_template": email_template,
            "sms_template": sms_template,
            "enabled": enabled,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        
        # Upsert configuration
        result = db.reminder_configurations.update_one(
            {"tenant_id": tenant_id},
            {"$set": config},
            upsert=True
        )
        
        if result.upserted_id:
            config["id"] = str(result.upserted_id)
        else:
            existing_config = db.reminder_configurations.find_one({"tenant_id": tenant_id})
            config["id"] = str(existing_config["_id"])
        
        return config
    
    @staticmethod
    def get_reminder_configuration(tenant_id: str) -> Optional[Dict]:
        """
        Get automated reminder configuration
        
        Returns:
            Dict with reminder configuration or None
        """
        db = Database.get_db()
        
        config = db.reminder_configurations.find_one({"tenant_id": tenant_id})
        
        if config:
            config["id"] = str(config["_id"])
            del config["_id"]
        
        return config
    
    @staticmethod
    def send_reminder(
        tenant_id: str,
        invoice_id: str,
        reminder_type: str = "email",
        custom_message: Optional[str] = None
    ) -> Dict:
        """
        Send a payment reminder for an invoice
        
        Returns:
            Dict with reminder sent data
        """
        db = Database.get_db()
        
        # Get invoice and client information
        invoice = db.invoices.find_one({
            "_id": ObjectId(invoice_id),
            "tenant_id": tenant_id
        })
        
        if not invoice:
            raise NotFoundException("Invoice not found")
        
        client = db.clients.find_one({"_id": ObjectId(invoice["client_id"])})
        if not client:
            raise NotFoundException("Client not found")
        
        # Create reminder record
        reminder = {
            "tenant_id": tenant_id,
            "invoice_id": invoice_id,
            "client_id": invoice["client_id"],
            "reminder_type": reminder_type,
            "message": custom_message or f"Payment reminder for invoice {invoice.get('invoice_number')}",
            "sent_to": client.get("email") if reminder_type == "email" else client.get("phone"),
            "sent_at": datetime.utcnow(),
            "status": "sent"  # In production, this would be updated based on actual sending
        }
        
        result = db.payment_reminders.insert_one(reminder)
        
        # Update invoice with reminder sent date
        db.invoices.update_one(
            {"_id": ObjectId(invoice_id)},
            {
                "$set": {"last_reminder_sent": datetime.utcnow()},
                "$inc": {"reminder_count": 1}
            }
        )
        
        # Add collection note
        AgingReportsService.add_collection_note(
            tenant_id=tenant_id,
            invoice_id=invoice_id,
            note_text=f"Payment reminder sent via {reminder_type}",
            note_type="reminder",
            created_by="system"
        )
        
        reminder["id"] = str(result.inserted_id)
        del reminder["_id"]
        
        return reminder
    
    @staticmethod
    def get_reminder_history(
        tenant_id: str,
        client_id: Optional[str] = None,
        invoice_id: Optional[str] = None,
        limit: int = 50
    ) -> List[Dict]:
        """
        Get payment reminder history
        
        Returns:
            List of reminder records
        """
        db = Database.get_db()
        
        query = {"tenant_id": tenant_id}
        
        if client_id:
            query["client_id"] = client_id
        
        if invoice_id:
            query["invoice_id"] = invoice_id
        
        reminders = list(db.payment_reminders.find(query)
                        .sort("sent_at", -1)
                        .limit(limit))
        
        # Convert ObjectId to string and add invoice/client details
        for reminder in reminders:
            reminder["id"] = str(reminder["_id"])
            del reminder["_id"]
            
            # Get invoice details
            if reminder.get("invoice_id"):
                invoice = db.invoices.find_one({"_id": ObjectId(reminder["invoice_id"])})
                if invoice:
                    reminder["invoice_number"] = invoice.get("invoice_number")
                    reminder["invoice_amount"] = invoice.get("total")
            
            # Get client details
            if reminder.get("client_id"):
                client = db.clients.find_one({"_id": ObjectId(reminder["client_id"])})
                if client:
                    reminder["client_name"] = client.get("name")
        
        return reminders
    
    @staticmethod
    def export_aging_report(
        tenant_id: str,
        format_type: str = "excel",
        as_of_date: Optional[str] = None,
        client_id: Optional[str] = None,
        include_zero_balances: bool = False
    ) -> BytesIO:
        """
        Export aging report to Excel or CSV
        
        Returns:
            BytesIO buffer with exported data
        """
        # Get aging data
        aging_data = AgingReportsService.get_detailed_ar_aging(
            tenant_id=tenant_id,
            as_of_date=as_of_date,
            client_id=client_id,
            include_zero_balances=include_zero_balances
        )
        
        if format_type.lower() == "excel":
            return AgingReportsService._export_aging_excel(aging_data)
        else:
            return AgingReportsService._export_aging_csv(aging_data)
    
    @staticmethod
    def _export_aging_excel(aging_data: Dict) -> BytesIO:
        """Export aging data to Excel format"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        buffer = BytesIO()
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Aging Summary"
        
        # Title
        ws_summary['A1'] = "Accounts Receivable Aging Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary['A2'] = f"As of: {aging_data['as_of_date']}"
        
        # Summary statistics
        row = 4
        ws_summary[f'A{row}'] = "Summary Statistics"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        summary_stats = [
            ("Total Outstanding", aging_data['total_outstanding']),
            ("Total Invoices", aging_data['summary']['total_invoices']),
            ("Total Clients", aging_data['summary']['total_clients']),
            ("Average Days Outstanding", aging_data['summary']['average_days_outstanding']),
            ("Largest Outstanding", aging_data['summary']['largest_outstanding']),
            ("Oldest Invoice (Days)", aging_data['summary']['oldest_invoice_days'])
        ]
        
        for label, value in summary_stats:
            ws_summary[f'A{row}'] = label
            ws_summary[f'B{row}'] = value
            if isinstance(value, float) and 'Outstanding' in label:
                ws_summary[f'B{row}'].number_format = '$#,##0.00'
            row += 1
        
        # Aging buckets
        row += 2
        ws_summary[f'A{row}'] = "Aging Buckets"
        ws_summary[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers
        headers = ['Bucket', 'Amount', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        row += 1
        
        # Bucket data
        total_amount = aging_data['total_outstanding']
        for bucket_key, bucket_data in aging_data['buckets'].items():
            percentage = (bucket_data['amount'] / total_amount * 100) if total_amount > 0 else 0
            
            ws_summary.cell(row=row, column=1, value=bucket_data['label'])
            ws_summary.cell(row=row, column=2, value=bucket_data['amount']).number_format = '$#,##0.00'
            ws_summary.cell(row=row, column=3, value=bucket_data['count'])
            ws_summary.cell(row=row, column=4, value=percentage/100).number_format = '0.0%'
            row += 1
        
        # Detail sheet
        ws_detail = wb.create_sheet("Invoice Details")
        
        # Headers
        detail_headers = [
            'Client Name', 'Invoice Number', 'Invoice Date', 'Due Date',
            'Total Amount', 'Amount Paid', 'Amount Due', 'Days Outstanding', 'Status'
        ]
        
        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Detail data
        row = 2
        for bucket_data in aging_data['buckets'].values():
            for invoice in bucket_data['invoices']:
                ws_detail.cell(row=row, column=1, value=invoice['client_name'])
                ws_detail.cell(row=row, column=2, value=invoice['invoice_number'])
                ws_detail.cell(row=row, column=3, value=invoice['invoice_date'])
                ws_detail.cell(row=row, column=4, value=invoice['due_date'])
                ws_detail.cell(row=row, column=5, value=invoice['total']).number_format = '$#,##0.00'
                ws_detail.cell(row=row, column=6, value=invoice['amount_paid']).number_format = '$#,##0.00'
                ws_detail.cell(row=row, column=7, value=invoice['amount_due']).number_format = '$#,##0.00'
                ws_detail.cell(row=row, column=8, value=invoice['days_outstanding'])
                ws_detail.cell(row=row, column=9, value=invoice['status'])
                row += 1
        
        # Auto-adjust column widths
        for ws in [ws_summary, ws_detail]:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def _export_aging_csv(aging_data: Dict) -> BytesIO:
        """Export aging data to CSV format"""
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Write summary
        writer.writerow(['Accounts Receivable Aging Report'])
        writer.writerow([f'As of: {aging_data["as_of_date"]}'])
        writer.writerow([])
        
        # Write aging buckets
        writer.writerow(['Aging Bucket', 'Amount', 'Count'])
        for bucket_key, bucket_data in aging_data['buckets'].items():
            writer.writerow([bucket_data['label'], bucket_data['amount'], bucket_data['count']])
        
        writer.writerow([])
        
        # Write invoice details
        writer.writerow([
            'Client Name', 'Invoice Number', 'Invoice Date', 'Due Date',
            'Total Amount', 'Amount Paid', 'Amount Due', 'Days Outstanding', 'Status'
        ])
        
        for bucket_data in aging_data['buckets'].values():
            for invoice in bucket_data['invoices']:
                writer.writerow([
                    invoice['client_name'],
                    invoice['invoice_number'],
                    invoice['invoice_date'],
                    invoice['due_date'],
                    invoice['total'],
                    invoice['amount_paid'],
                    invoice['amount_due'],
                    invoice['days_outstanding'],
                    invoice['status']
                ])
        
        # Convert to BytesIO
        buffer = BytesIO()
        buffer.write(output.getvalue().encode('utf-8'))
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def _get_client_payment_history(
        db, 
        tenant_id: str, 
        client_id: str, 
        limit: int = 10
    ) -> List[Dict]:
        """Get recent payment history for a client"""
        # Get invoices for this client
        invoices = list(db.invoices.find({
            "tenant_id": tenant_id,
            "client_id": client_id
        }))
        
        invoice_ids = [str(inv["_id"]) for inv in invoices]
        
        # Get payments for these invoices
        payments = list(db.payments.find({
            "tenant_id": tenant_id,
            "invoice_id": {"$in": invoice_ids}
        }).sort("payment_date", -1).limit(limit))
        
        # Convert ObjectId to string and add invoice details
        for payment in payments:
            payment["id"] = str(payment["_id"])
            del payment["_id"]
            
            # Get invoice details
            invoice = next((inv for inv in invoices if str(inv["_id"]) == payment["invoice_id"]), None)
            if invoice:
                payment["invoice_number"] = invoice.get("invoice_number")
        
        return payments
    
    @staticmethod
    def _get_collection_notes(
        db, 
        tenant_id: str, 
        client_id: str
    ) -> List[Dict]:
        """Get collection notes for a client"""
        notes = list(db.collection_notes.find({
            "tenant_id": tenant_id,
            "client_id": client_id
        }).sort("created_at", -1).limit(20))
        
        # Convert ObjectId to string
        for note in notes:
            note["id"] = str(note["_id"])
            del note["_id"]
        
        return notes


# Singleton instance
aging_reports_service = AgingReportsService()