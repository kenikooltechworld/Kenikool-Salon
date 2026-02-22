"""
Report generation service for creating and exporting analytics reports
"""
import logging
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
import json
import csv
from io import StringIO, BytesIO

logger = logging.getLogger(__name__)


class ReportGenerationService:
    """Service for generating and exporting analytics reports"""

    def __init__(self):
        """Initialize report generation service"""
        self.reports = {}
        self.schedules = {}

    async def generate_custom_report(
        self,
        tenant_id: str,
        report_name: str,
        metrics: List[str],
        filters: List[Dict[str, Any]],
        date_range: Dict[str, datetime]
    ) -> Dict[str, Any]:
        """Generate custom report"""
        try:
            report_id = f"report_{datetime.utcnow().timestamp()}"
            
            report_data = {
                "report_id": report_id,
                "tenant_id": tenant_id,
                "name": report_name,
                "metrics": metrics,
                "filters": filters,
                "date_range": date_range,
                "created_at": datetime.utcnow().isoformat(),
                "status": "completed",
                "data": self._generate_sample_data(metrics)
            }
            
            self.reports[report_id] = report_data
            return report_data
        except Exception as e:
            logger.error(f"Error generating custom report: {e}")
            raise

    async def export_data(
        self,
        tenant_id: str,
        format: str,
        data: List[Dict[str, Any]],
        filename: str = "export"
    ) -> Dict[str, Any]:
        """Export data in specified format"""
        try:
            export_id = f"export_{datetime.utcnow().timestamp()}"
            
            if format == "csv":
                content = self._export_csv(data)
                file_extension = "csv"
            elif format == "json":
                content = self._export_json(data)
                file_extension = "json"
            elif format == "excel":
                content = self._export_excel(data)
                file_extension = "xlsx"
            elif format == "pdf":
                content = self._export_pdf(data)
                file_extension = "pdf"
            else:
                raise ValueError(f"Unsupported format: {format}")
            
            return {
                "export_id": export_id,
                "status": "completed",
                "format": format,
                "filename": f"{filename}.{file_extension}",
                "size_bytes": len(content) if isinstance(content, (str, bytes)) else 0,
                "created_at": datetime.utcnow().isoformat(),
                "expires_at": (datetime.utcnow() + timedelta(days=7)).isoformat()
            }
        except Exception as e:
            logger.error(f"Error exporting data: {e}")
            raise

    async def schedule_report(
        self,
        tenant_id: str,
        report_id: str,
        schedule: str,
        recipients: List[str]
    ) -> Dict[str, Any]:
        """Schedule report for automated generation and delivery"""
        try:
            schedule_id = f"schedule_{datetime.utcnow().timestamp()}"
            
            schedule_data = {
                "schedule_id": schedule_id,
                "tenant_id": tenant_id,
                "report_id": report_id,
                "schedule": schedule,
                "recipients": recipients,
                "status": "active",
                "created_at": datetime.utcnow().isoformat(),
                "next_run": self._calculate_next_run(schedule)
            }
            
            self.schedules[schedule_id] = schedule_data
            logger.info(f"Report {report_id} scheduled with ID {schedule_id}")
            return schedule_data
        except Exception as e:
            logger.error(f"Error scheduling report: {e}")
            raise

    def _calculate_next_run(self, schedule: str) -> str:
        """Calculate next run time based on schedule"""
        now = datetime.utcnow()
        if schedule == "daily":
            next_run = now + timedelta(days=1)
        elif schedule == "weekly":
            next_run = now + timedelta(weeks=1)
        elif schedule == "monthly":
            next_run = now + timedelta(days=30)
        else:
            next_run = now + timedelta(days=1)
        return next_run.isoformat()

    def _generate_sample_data(self, metrics: List[str]) -> List[Dict[str, Any]]:
        """Generate sample data for report"""
        data = []
        for i in range(30):
            row = {"date": (datetime.utcnow() - timedelta(days=30-i)).isoformat()}
            for metric in metrics:
                if metric == "revenue":
                    row[metric] = 5000 + (i * 100)
                elif metric == "bookings":
                    row[metric] = 50 + i
                elif metric == "clients":
                    row[metric] = 100 + (i * 2)
                else:
                    row[metric] = i * 10
            data.append(row)
        return data

    def _export_csv(self, data: List[Dict[str, Any]]) -> str:
        """Export data as CSV"""
        if not data:
            return ""
        
        output = StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        return output.getvalue()

    def _export_json(self, data: List[Dict[str, Any]]) -> str:
        """Export data as JSON"""
        return json.dumps(data, indent=2, default=str)

    def _export_excel(self, data: List[Dict[str, Any]]) -> bytes:
        """Export data as Excel"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            if data:
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            
            output = BytesIO()
            wb.save(output)
            return output.getvalue()
        except ImportError:
            logger.warning("openpyxl not available, returning empty bytes")
            return b""

    def _export_pdf(self, data: List[Dict[str, Any]]) -> bytes:
        """Export data as PDF"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
            from reportlab.lib import colors
            
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter)
            
            if data:
                headers = list(data[0].keys())
                table_data = [headers]
                for row in data:
                    table_data.append([str(row.get(h, "")) for h in headers])
                
                table = Table(table_data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                doc.build([table])
            
            return output.getvalue()
        except ImportError:
            logger.warning("reportlab not available, returning empty bytes")
            return b""


from datetime import timedelta
