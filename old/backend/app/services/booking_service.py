"""
Booking service - Business logic layer
"""
from bson import ObjectId
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import logging

from app.database import Database
from app.api.exceptions import NotFoundException, BadRequestException

logger = logging.getLogger(__name__)


class BookingService:
    """Booking service for handling booking business logic"""
    
    # ========================================================================
    # Location Validation Methods
    # ========================================================================
    
    @staticmethod
    def validate_location_exists(location_id: str, tenant_id: str) -> bool:
        """
        Validate that location exists and belongs to tenant
        
        Requirements: 6.1.2
        
        Args:
            location_id: Location ID to validate
            tenant_id: Tenant ID
            
        Returns:
            True if location exists, raises exception otherwise
        """
        db = Database.get_db()
        
        location = db.locations.find_one({
            "_id": ObjectId(location_id),
            "tenant_id": tenant_id
        })
        
        if not location:
            raise NotFoundException(f"Location not found: {location_id}")
        
        return True
    
    @staticmethod
    def validate_staff_at_location(staff_id: str, location_id: str, tenant_id: str) -> bool:
        """
        Validate that staff member is assigned to location
        
        Requirements: 6.2.1
        
        Args:
            staff_id: Staff ID to validate
            location_id: Location ID
            tenant_id: Tenant ID
            
        Returns:
            True if staff is at location, raises exception otherwise
        """
        db = Database.get_db()
        
        staff = db.stylists.find_one({
            "_id": ObjectId(staff_id),
            "tenant_id": tenant_id
        })
        
        if not staff:
            raise NotFoundException(f"Staff member not found: {staff_id}")
        
        # Check if staff is assigned to location
        assigned_locations = staff.get("assigned_locations", [])
        if location_id not in assigned_locations:
            raise BadRequestException(f"Staff member is not assigned to this location")
        
        return True
    
    @staticmethod
    def validate_service_at_location(service_id: str, location_id: str, tenant_id: str) -> bool:
        """
        Validate that service is offered at location
        
        Requirements: 6.2.3
        
        Args:
            service_id: Service ID to validate
            location_id: Location ID
            tenant_id: Tenant ID
            
        Returns:
            True if service is at location, raises exception otherwise
        """
        db = Database.get_db()
        
        service = db.services.find_one({
            "_id": ObjectId(service_id),
            "tenant_id": tenant_id
        })
        
        if not service:
            raise NotFoundException(f"Service not found: {service_id}")
        
        # Check if service is offered at location
        offered_locations = service.get("offered_locations", [])
        if location_id not in offered_locations:
            raise BadRequestException(f"Service is not offered at this location")
        
        return True
    
    @staticmethod
    def check_double_booking(
        staff_id: str,
        location_id: str,
        start_time: datetime,
        end_time: datetime,
        tenant_id: str,
        exclude_booking_id: Optional[str] = None
    ) -> bool:
        """
        Check for double-booking conflicts at location
        
        Requirements: 6.2.4
        
        Args:
            staff_id: Staff ID
            location_id: Location ID
            start_time: Booking start time
            end_time: Booking end time
            tenant_id: Tenant ID
            exclude_booking_id: Booking ID to exclude from check (for updates)
            
        Returns:
            True if no conflicts, raises exception if conflict found
        """
        db = Database.get_db()
        
        # Build query for conflicting bookings
        query = {
            "tenant_id": tenant_id,
            "stylist_id": staff_id,
            "location_id": location_id,
            "status": {"$in": ["pending", "confirmed"]},
            "booking_date": {
                "$lt": end_time,
                "$gte": start_time
            }
        }
        
        # Exclude current booking if updating
        if exclude_booking_id:
            query["_id"] = {"$ne": ObjectId(exclude_booking_id)}
        
        conflict = db.bookings.find_one(query)
        
        if conflict:
            raise BadRequestException(
                f"Staff member has a conflicting booking at this location during the requested time"
            )
        
        return True
    
    @staticmethod
    def get_location_specific_pricing(
        service_id: str,
        location_id: str,
        tenant_id: str
    ) -> float:
        """
        Get location-specific service pricing
        
        Requirements: 6.2.5
        
        Args:
            service_id: Service ID
            location_id: Location ID
            tenant_id: Tenant ID
            
        Returns:
            Location-specific price or default service price
        """
        db = Database.get_db()
        
        # Check for location-specific pricing
        location_pricing = db.location_service_pricing.find_one({
            "service_id": service_id,
            "location_id": location_id,
            "tenant_id": tenant_id
        })
        
        if location_pricing:
            return location_pricing.get("price", 0.0)
        
        # Fall back to default service pricing
        service = db.services.find_one({
            "_id": ObjectId(service_id),
            "tenant_id": tenant_id
        })
        
        if service:
            return service.get("price", 0.0)
        
        return 0.0
    
    @staticmethod
    def filter_bookings_by_location(
        tenant_id: str,
        location_id: Optional[str] = None,
        location_ids: Optional[List[str]] = None,
        status: Optional[str] = None,
        stylist_id: Optional[str] = None,
        client_id: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        offset: int = 0,
        limit: int = 20
    ) -> List[Dict]:
        """
        Filter bookings by location
        
        Requirements: 6.2.6
        
        Args:
            tenant_id: Tenant ID
            location_id: Single location ID to filter by
            location_ids: Multiple location IDs to filter by
            status: Booking status filter
            stylist_id: Stylist ID filter
            client_id: Client ID filter
            date_from: Start date filter
            date_to: End date filter
            offset: Pagination offset
            limit: Pagination limit
            
        Returns:
            List of bookings matching filters
        """
        db = Database.get_db()
        
        # Build query
        query = {"tenant_id": tenant_id}
        
        # Add location filter
        if location_id:
            query["location_id"] = location_id
        elif location_ids:
            query["location_id"] = {"$in": location_ids}
        
        # Add other filters
        if status:
            query["status"] = status
        if stylist_id:
            query["stylist_id"] = stylist_id
        if client_id:
            query["client_id"] = client_id
        
        # Add date filters
        if date_from or date_to:
            query["booking_date"] = {}
            if date_from:
                try:
                    date_from_dt = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
                except ValueError:
                    date_from_dt = datetime.strptime(date_from, '%Y-%m-%d')
                query["booking_date"]["$gte"] = date_from_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            if date_to:
                try:
                    date_to_dt = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
                except ValueError:
                    date_to_dt = datetime.strptime(date_to, '%Y-%m-%d')
                query["booking_date"]["$lte"] = date_to_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        bookings = list(db.bookings.find(query).sort("booking_date", -1).skip(offset).limit(limit))
        
        return [BookingService._format_booking_response(b) for b in bookings]
    
    @staticmethod
    def get_available_staff_for_location(
        location_id: str,
        tenant_id: str,
        service_id: Optional[str] = None,
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None
    ) -> List[Dict]:
        """
        Get staff available at location for service during time
        
        Requirements: 6.2.7
        
        Args:
            location_id: Location ID
            tenant_id: Tenant ID
            service_id: Optional service ID to filter by
            start_time: Optional start time for availability check
            end_time: Optional end time for availability check
            
        Returns:
            List of available staff
        """
        db = Database.get_db()
        
        # Get all staff assigned to location
        query = {
            "tenant_id": tenant_id,
            "assigned_locations": location_id,
            "is_active": True
        }
        
        staff_list = list(db.stylists.find(query))
        
        # Filter by service if provided
        if service_id:
            staff_list = [
                s for s in staff_list
                if service_id in s.get("services", [])
            ]
        
        # Filter by availability if time provided
        if start_time and end_time:
            available_staff = []
            for staff in staff_list:
                # Check for conflicts
                conflict = db.bookings.find_one({
                    "tenant_id": tenant_id,
                    "stylist_id": str(staff["_id"]),
                    "location_id": location_id,
                    "status": {"$in": ["pending", "confirmed"]},
                    "booking_date": {
                        "$lt": end_time,
                        "$gte": start_time
                    }
                })
                
                if not conflict:
                    available_staff.append(staff)
            
            staff_list = available_staff
        
        return staff_list
    
    @staticmethod
    def get_available_services_for_location(
        location_id: str,
        tenant_id: str
    ) -> List[Dict]:
        """
        Get services offered at location
        
        Requirements: 6.2.7
        
        Args:
            location_id: Location ID
            tenant_id: Tenant ID
            
        Returns:
            List of services offered at location
        """
        db = Database.get_db()
        
        # Get all services offered at location
        services = list(db.services.find({
            "tenant_id": tenant_id,
            "offered_locations": location_id,
            "is_active": True
        }))
        
        return services
    
    @staticmethod
    def get_bookings(
        tenant_id: str,
        status: Optional[str] = None,
        stylist_id: Optional[str] = None,
        client_id: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        booking_type: Optional[str] = None,
        group_subtype: Optional[str] = None,
        offset: int = 0,
        limit: int = 20
    ) -> List[Dict]:
        """
        Get list of bookings for tenant with filtering and pagination
        
        Returns:
            List of booking dicts
        """
        db = Database.get_db()
        
        # Build query - convert tenant_id to ObjectId
        query = {"tenant_id": ObjectId(tenant_id)}
        if status:
            query["status"] = status
        if stylist_id:
            query["stylist_id"] = stylist_id
        if client_id:
            query["client_id"] = client_id
        if booking_type:
            query["booking_type"] = booking_type
        if group_subtype:
            query["group_subtype"] = group_subtype
        if date_from or date_to:
            query["booking_date"] = {}
            if date_from:
                # Start of the day
                # Handle both ISO format with timezone and simple date format
                try:
                    date_from_dt = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
                except ValueError:
                    # If no timezone, parse as simple date
                    date_from_dt = datetime.strptime(date_from, '%Y-%m-%d')
                query["booking_date"]["$gte"] = date_from_dt.replace(hour=0, minute=0, second=0, microsecond=0)
            if date_to:
                # End of the day (23:59:59.999999)
                try:
                    date_to_dt = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
                except ValueError:
                    # If no timezone, parse as simple date
                    date_to_dt = datetime.strptime(date_to, '%Y-%m-%d')
                query["booking_date"]["$lte"] = date_to_dt.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        bookings = list(db.bookings.find(query).sort("booking_date", -1).skip(offset).limit(limit))
        
        return [BookingService._format_booking_response(b) for b in bookings]
    
    @staticmethod
    def get_booking(booking_id: str, tenant_id: str) -> Dict:
        """
        Get single booking by ID
        
        Returns:
            Dict with booking data
        """
        db = Database.get_db()
        
        booking_doc = db.bookings.find_one({
            "_id": ObjectId(booking_id),
            "tenant_id": tenant_id
        })
        
        if booking_doc is None:
            raise NotFoundException("Booking not found")
        
        return BookingService._format_booking_response(booking_doc)
    
    @staticmethod
    async def create_booking(
        tenant_id: str,
        location_id: str,
        client_name: str,
        client_phone: str,
        service_id: str,
        stylist_id: str,
        booking_date: str,
        client_email: Optional[str] = None,
        notes: Optional[str] = None,
        variant_id: Optional[str] = None
    ) -> Dict:
        """
        Create a new booking
        
        Requirements: 3.4, 6.1, 6.2
        
        Returns:
            Dict with created booking data
        """
        db = Database.get_db()
        
        # Validate location exists
        BookingService.validate_location_exists(location_id, tenant_id)
        
        # Validate staff is at location
        BookingService.validate_staff_at_location(stylist_id, location_id, tenant_id)
        
        # Validate service is at location
        BookingService.validate_service_at_location(service_id, location_id, tenant_id)
        
        # Parse booking date
        try:
            booking_datetime = datetime.fromisoformat(booking_date.replace('Z', '+00:00'))
        except ValueError:
            raise BadRequestException("Invalid booking date format")
        
        # Get service details
        service = db.services.find_one({
            "_id": ObjectId(service_id),
            "tenant_id": tenant_id
        })
        
        if not service:
            raise NotFoundException("Service not found")
        
        if not service.get("is_active"):
            raise BadRequestException("Service is not available")
        
        # Check prerequisite services
        prerequisite_services = service.get("prerequisite_services", [])
        if prerequisite_services:
            # Check if client has completed all prerequisite services
            for prereq_id in prerequisite_services:
                completed_prereq = db.bookings.find_one({
                    "client_id": client_id if 'client_id' in locals() else None,
                    "service_id": prereq_id,
                    "tenant_id": tenant_id,
                    "status": "completed"
                })
                
                if not completed_prereq:
                    prereq_service = db.services.find_one({"_id": ObjectId(prereq_id)})
                    prereq_name = prereq_service.get("name", "Unknown") if prereq_service else "Unknown"
                    raise BadRequestException(f"Prerequisite service '{prereq_name}' must be completed first")
        
        # Get variant details if specified
        variant = None
        service_price = BookingService.get_location_specific_pricing(service_id, location_id, tenant_id)
        duration_minutes = service.get("duration_minutes", 30)
        variant_name = None
        
        if variant_id:
            variant = db.service_variants.find_one({
                "_id": ObjectId(variant_id),
                "service_id": service_id,
                "tenant_id": tenant_id
            })
            
            if not variant:
                raise NotFoundException("Variant not found")
            
            if not variant.get("is_active"):
                raise BadRequestException("Variant is not available")
            
            # Use variant price and duration
            service_price = variant.get("final_price", service_price)
            duration_minutes = variant.get("final_duration", duration_minutes)
            variant_name = variant.get("name")
        
        # Get stylist details
        stylist = db.stylists.find_one({
            "_id": ObjectId(stylist_id),
            "tenant_id": tenant_id
        })
        
        if stylist is None:
            raise NotFoundException("Stylist not found")
        
        if not stylist.get("is_active"):
            raise BadRequestException("Stylist is not available")
        
        # Validate booking rules
        from app.services.booking_rules_service import get_booking_rules_service
        rules_service = get_booking_rules_service(db)
        
        # Check booking rules
        is_valid, rules_error = rules_service.validate_booking_rules(
            service_id, booking_datetime, tenant_id, stylist_id
        )
        if not is_valid:
            raise BadRequestException(f"Booking rules violation: {rules_error}")
        
        # Check service availability
        is_available, availability_error = rules_service.check_service_availability(
            service_id, booking_datetime, tenant_id
        )
        if not is_available:
            raise BadRequestException(f"Service not available: {availability_error}")
        
        # Check capacity
        has_capacity, available_slots = rules_service.check_capacity(
            service_id, booking_datetime, tenant_id
        )
        if not has_capacity:
            raise BadRequestException(f"Service is fully booked (no available slots)")
        
        # Check double-booking at location
        BookingService.check_double_booking(
            stylist_id, location_id, booking_datetime,
            booking_datetime + timedelta(minutes=duration_minutes),
            tenant_id
        )
        
        # Check availability
        from app.services.availability_service import is_slot_available, get_existing_bookings
        existing_bookings = await get_existing_bookings(stylist_id, booking_datetime, tenant_id)
        if not is_slot_available(booking_datetime, duration_minutes, existing_bookings):
            raise BadRequestException("Time slot is not available")
        
        # Find or create client
        client = db.clients.find_one({
            "tenant_id": tenant_id,
            "phone": client_phone
        })
        
        if client is not None:
            client_id = str(client["_id"])
            # Update client info if needed
            db.clients.update_one(
                {"_id": client["_id"]},
                {
                    "$set": {
                        "name": client_name,
                        "email": client_email,
                        "updated_at": datetime.utcnow()
                    }
                }
            )
        else:
            # Create new client
            client_data = {
                "tenant_id": tenant_id,
                "name": client_name,
                "phone": client_phone,
                "email": client_email,
                "total_visits": 0,
                "total_spent": 0.0,
                "created_at": datetime.utcnow(),
                "updated_at": datetime.utcnow()
            }
            result = db.clients.insert_one(client_data)
            client_id = str(result.inserted_id)
        
        # Get location name
        location = db.locations.find_one({
            "_id": ObjectId(location_id),
            "tenant_id": tenant_id
        })
        location_name = location.get("name", "Unknown Location") if location else "Unknown Location"
        
        # Create booking
        booking_data = {
            "tenant_id": tenant_id,
            "location_id": location_id,
            "client_id": client_id,
            "client_name": client_name,
            "client_phone": client_phone,
            "service_id": service_id,
            "service_name": service.get("name", "Unknown Service"),
            "service_price": service_price,
            "stylist_id": stylist_id,
            "stylist_name": stylist.get("name", "Unknown Stylist"),
            "booking_date": booking_datetime,
            "duration_minutes": duration_minutes,
            "status": "pending",
            "payment_status": "unpaid",
            "payment_method": None,
            "notes": notes,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        
        # Add variant info if present
        if variant_id:
            booking_data["variant_id"] = variant_id
            booking_data["variant_name"] = variant_name
        
        result = db.bookings.insert_one(booking_data)
        booking_id = str(result.inserted_id)
        
        logger.info(f"Booking created: {booking_id} for tenant: {tenant_id} at location: {location_id}")
        
        # Create notification for new booking
        try:
            from app.services.notification_service import NotificationService
            notification = await NotificationService.create_booking_notification(
                db=db,
                tenant_id=tenant_id,
                booking_id=booking_id,
                client_name=client_name,
                service_name=service.get("name", "Unknown Service"),
                booking_date=booking_datetime.strftime("%B %d, %Y at %I:%M %p")
            )
            logger.info(f"Booking notification created: {notification.get('_id')} for booking: {booking_id}")
        except Exception as e:
            logger.error(f"Failed to create booking notification: {e}", exc_info=True)
        
        # Send confirmation message asynchronously
        try:
            from app.tasks.reminder_tasks import send_immediate_confirmation
            send_immediate_confirmation.delay(booking_id)
        except Exception as e:
            logger.warning(f"Failed to queue confirmation message: {e}")
        
        # Fetch created booking
        booking_doc = db.bookings.find_one({"_id": ObjectId(booking_id)})
        return BookingService._format_booking_response(booking_doc)
    
    @staticmethod
    def update_booking_status(
        booking_id: str,
        tenant_id: str,
        status: str
    ) -> Dict:
        """
        Update booking status
        
        Returns:
            Dict with updated booking data
        """
        db = Database.get_db()
        
        # Validate status
        valid_statuses = ["pending", "confirmed", "completed", "cancelled", "no_show"]
        if status not in valid_statuses:
            raise BadRequestException(f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
        
        # Get booking
        booking_doc = db.bookings.find_one({
            "_id": ObjectId(booking_id),
            "tenant_id": tenant_id
        })
        
        if booking_doc is None:
            raise NotFoundException("Booking not found")
        
        # Update booking
        update_data = {
            "status": status,
            "updated_at": datetime.utcnow()
        }
        
        if status == "confirmed" and not booking_doc.get("confirmed_at"):
            update_data["confirmed_at"] = datetime.utcnow()
            
            # Send confirmation message if not already sent
            if not booking_doc.get("confirmation_sent"):
                try:
                    from app.tasks.reminder_tasks import send_immediate_confirmation
                    send_immediate_confirmation.delay(booking_id)
                except Exception as e:
                    logger.warning(f"Failed to queue confirmation message: {e}")
        elif status == "completed" and not booking_doc.get("completed_at"):
            update_data["completed_at"] = datetime.utcnow()
            
            # Redeem package credit if booking was paid via package credit
            if booking_doc.get("paid_via_package_credit") and booking_doc.get("package_credit_id"):
                try:
                    BookingService.redeem_package_credit_for_booking(
                        booking_id=booking_id,
                        tenant_id=tenant_id,
                        staff_id=booking_doc.get("stylist_id", "system")
                    )
                except Exception as e:
                    logger.error(f"Failed to redeem package credit for booking {booking_id}: {e}")
                    # Don't fail booking completion if credit redemption fails
            
            # Update client stats
            db.clients.update_one(
                {"_id": ObjectId(booking_doc["client_id"])},
                {
                    "$inc": {
                        "total_visits": 1,
                        "total_spent": booking_doc["service_price"] if not booking_doc.get("paid_via_package_credit") else 0
                    },
                    "$set": {
                        "last_visit_date": datetime.utcnow(),
                        "last_activity_date": datetime.utcnow()
                    }
                }
            )
            
            # Check if this is the first booking and complete referral if applicable
            client_doc = db.clients.find_one({"_id": ObjectId(booking_doc["client_id"])})
            if client_doc and client_doc.get("total_visits", 0) == 1:
                try:
                    from app.services.referral_service import ReferralService
                    from app.config import settings
                    
                    # Get referral reward amount from tenant settings or use default
                    tenant = db.tenants.find_one({"_id": ObjectId(tenant_id)})
                    reward_amount = 1000  # Default reward amount
                    
                    if tenant and tenant.get("referral_settings"):
                        reward_amount = tenant["referral_settings"].get("reward_amount", 1000)
                    
                    # Complete the referral
                    ReferralService.complete_referral(
                        tenant_id=tenant_id,
                        referred_client_id=booking_doc["client_id"],
                        reward_amount=reward_amount
                    )
                    logger.info(f"Referral completed for client: {booking_doc['client_id']}, Reward: {reward_amount}")
                except Exception as e:
                    logger.warning(f"Failed to complete referral: {str(e)}")
                    # Don't fail booking completion if referral completion fails
            
            # Trigger analytics calculation for client
            try:
                from app.tasks.analytics_tasks import calculate_client_analytics
                calculate_client_analytics.delay(booking_doc["client_id"], tenant_id)
                logger.info(f"Queued analytics calculation for client: {booking_doc['client_id']}")
            except Exception as e:
                logger.warning(f"Failed to queue analytics calculation: {e}")
            
            # Track commission (synchronous, no Celery needed)
            try:
                from app.services.commission_service import track_commission
                
                # Get stylist commission settings
                stylist = db.stylists.find_one({"_id": ObjectId(booking_doc["stylist_id"])})
                if stylist and stylist.get("commission_value"):
                    track_commission(
                        booking_id=booking_id,
                        stylist_id=booking_doc["stylist_id"],
                        service_price=booking_doc["service_price"],
                        commission_rate=stylist.get("commission_value", 0),
                        commission_type=stylist.get("commission_type", "percentage"),
                        tenant_id=tenant_id,
                        db=db
                    )
                    logger.info(f"Commission tracked for booking: {booking_id}")
            except Exception as e:
                logger.error(f"Failed to track commission: {e}", exc_info=True)
            
            # Note: Inventory deduction and review requests require Celery worker
            # These will be queued but won't execute without Celery running
            try:
                from app.tasks.inventory_tasks import deduct_inventory_for_service
                deduct_inventory_for_service.delay(booking_id)
                logger.info(f"Inventory deduction queued for booking: {booking_id}")
            except Exception as e:
                logger.warning(f"Failed to queue inventory deduction (Celery may not be running): {e}")
            
            try:
                from app.tasks.review_tasks import send_review_request
                send_review_request.delay(booking_id)
                logger.info(f"Review request queued for booking: {booking_id}")
            except Exception as e:
                logger.warning(f"Failed to queue review request (Celery may not be running): {e}")
            
            # Update family loyalty points if this is a family booking
            if booking_doc.get("family_account_id"):
                try:
                    from app.tasks.booking_tasks import update_family_loyalty_points
                    update_family_loyalty_points.delay(booking_id)
                    logger.info(f"Family loyalty points update queued for booking: {booking_id}")
                except Exception as e:
                    logger.warning(f"Failed to queue family loyalty points update: {e}")
        elif status == "cancelled" and not booking_doc.get("cancelled_at"):
            update_data["cancelled_at"] = datetime.utcnow()
            
            # Release package credit if booking was paid via package credit
            if booking_doc.get("paid_via_package_credit") and booking_doc.get("package_credit_id"):
                try:
                    BookingService.release_package_credit_for_booking(
                        booking_id=booking_id,
                        tenant_id=tenant_id
                    )
                except Exception as e:
                    logger.error(f"Failed to release package credit for booking {booking_id}: {e}")
                    # Don't fail booking cancellation if credit release fails
            
            # Trigger waitlist notifications for cancelled booking
            try:
                from app.tasks.booking_tasks import sync_waitlist_notifications
                sync_waitlist_notifications.delay(booking_id)
                logger.info(f"Waitlist notifications queued for cancelled booking: {booking_id}")
            except Exception as e:
                logger.warning(f"Failed to queue waitlist notifications: {e}")
        
        db.bookings.update_one(
            {"_id": ObjectId(booking_id)},
            {"$set": update_data}
        )
        
        logger.info(f"Booking status updated: {booking_id} -> {status}")
        
        # Fetch updated booking
        booking_doc = db.bookings.find_one({"_id": ObjectId(booking_id)})
        return BookingService._format_booking_response(booking_doc)
    
    @staticmethod
    def update_payment_method(
        booking_id: str,
        tenant_id: str,
        payment_method: str
    ) -> Dict:
        """
        Update booking payment method
        
        Returns:
            Dict with updated booking data
        """
        db = Database.get_db()
        
        # Get booking
        booking_doc = db.bookings.find_one({
            "_id": ObjectId(booking_id),
            "tenant_id": tenant_id
        })
        
        if booking_doc is None:
            raise NotFoundException("Booking not found")
        
        # Update booking
        update_data = {
            "payment_method": payment_method,
            "updated_at": datetime.utcnow()
        }
        
        # If payment method is bank_transfer, set status to pending_payment
        if payment_method == "bank_transfer":
            update_data["payment_status"] = "pending_payment"
        
        db.bookings.update_one(
            {"_id": ObjectId(booking_id)},
            {"$set": update_data}
        )
        
        logger.info(f"Booking payment method updated: {booking_id} -> {payment_method}")
        
        # Fetch updated booking
        booking_doc = db.bookings.find_one({"_id": ObjectId(booking_id)})
        return BookingService._format_booking_response(booking_doc)
    
    @staticmethod
    def update_payment_status(
        booking_id: str,
        tenant_id: str,
        payment_status: str
    ) -> Dict:
        """
        Update booking payment status
        
        Returns:
            Dict with updated booking data
        """
        db = Database.get_db()
        
        # Get booking
        booking_doc = db.bookings.find_one({
            "_id": ObjectId(booking_id),
            "tenant_id": tenant_id
        })
        
        if booking_doc is None:
            raise NotFoundException("Booking not found")
        
        # Update booking
        update_data = {
            "payment_status": payment_status,
            "updated_at": datetime.utcnow()
        }
        
        # If payment is confirmed, record payment date
        if payment_status == "paid" and not booking_doc.get("paid_at"):
            update_data["paid_at"] = datetime.utcnow()
        
        db.bookings.update_one(
            {"_id": ObjectId(booking_id)},
            {"$set": update_data}
        )
        
        logger.info(f"Booking payment status updated: {booking_id} -> {payment_status}")
        
        # Fetch updated booking
        booking_doc = db.bookings.find_one({"_id": ObjectId(booking_id)})
        return BookingService._format_booking_response(booking_doc)
    
    @staticmethod
    def edit_booking(
        booking_id: str,
        tenant_id: str,
        updates: Dict,
        user_id: Optional[str] = None,
        user_name: Optional[str] = None
    ) -> Dict:
        """
        Edit an existing booking
        
        Requirements: 2.1, 2.2, 2.3, 2.4, 2.5, 2.6, 2.7
        
        Args:
            booking_id: Booking ID to edit
            tenant_id: Tenant ID
            updates: Dictionary of fields to update
            user_id: User performing the edit
            user_name: Name of user performing the edit
            
        Returns:
            Dict with updated booking data
        """
        db = Database.get_db()
        
        # Get existing booking
        booking_doc = db.bookings.find_one({
            "_id": ObjectId(booking_id),
            "tenant_id": ObjectId(tenant_id)
        })
        
        if not booking_doc:
            raise NotFoundException("Booking not found")
        
        # Track changes for audit log
        changes = {}
        update_data = {"updated_at": datetime.utcnow()}
        
        # Handle service change
        if "service_id" in updates and updates["service_id"]:
            service = db.services.find_one({
                "_id": ObjectId(updates["service_id"]),
                "tenant_id": ObjectId(tenant_id)
            })
            if not service:
                raise NotFoundException("Service not found")
            
            changes["service_id"] = {
                "old": str(booking_doc["service_id"]),
                "new": updates["service_id"]
            }
            update_data["service_id"] = ObjectId(updates["service_id"])
            update_data["service_name"] = service.get("name")
            update_data["service_price"] = service.get("price", 0.0)
            update_data["duration_minutes"] = service.get("duration_minutes", 30)
        
        # Handle stylist change
        if "stylist_id" in updates and updates["stylist_id"]:
            stylist = db.stylists.find_one({
                "_id": ObjectId(updates["stylist_id"]),
                "tenant_id": ObjectId(tenant_id)
            })
            if not stylist:
                raise NotFoundException("Stylist not found")
            
            # Check stylist availability if date is also being changed or use existing date
            booking_date = updates.get("booking_date", booking_doc["booking_date"])
            if isinstance(booking_date, str):
                try:
                    booking_datetime = datetime.fromisoformat(booking_date.replace('Z', '+00:00'))
                except ValueError:
                    booking_datetime = datetime.strptime(booking_date, '%Y-%m-%d')
            else:
                booking_datetime = booking_date
            
            # TODO: Check for conflicts with new stylist
            # This will be implemented in Task 1.5
            
            changes["stylist_id"] = {
                "old": str(booking_doc["stylist_id"]),
                "new": updates["stylist_id"]
            }
            update_data["stylist_id"] = ObjectId(updates["stylist_id"])
            update_data["stylist_name"] = stylist.get("name")
        
        # Handle date/time change
        if "booking_date" in updates and updates["booking_date"]:
            try:
                new_datetime = datetime.fromisoformat(updates["booking_date"].replace('Z', '+00:00'))
            except ValueError:
                new_datetime = datetime.strptime(updates["booking_date"], '%Y-%m-%d')
            
            changes["booking_date"] = {
                "old": booking_doc["booking_date"].isoformat() if isinstance(booking_doc["booking_date"], datetime) else str(booking_doc["booking_date"]),
                "new": new_datetime.isoformat()
            }
            update_data["booking_date"] = new_datetime
        
        # Handle client info changes
        for field in ["client_name", "client_phone", "client_email", "notes", "group_subtype"]:
            if field in updates and updates[field] is not None:
                if booking_doc.get(field) != updates[field]:
                    changes[field] = {
                        "old": booking_doc.get(field),
                        "new": updates[field]
                    }
                    update_data[field] = updates[field]
        
        # Update booking
        db.bookings.update_one(
            {"_id": ObjectId(booking_id)},
            {"$set": update_data}
        )
        
        # Create audit log
        if changes:
            from app.schemas.booking_audit import AuditActionType
            db.booking_audit_logs.insert_one({
                "booking_id": booking_id,
                "tenant_id": ObjectId(tenant_id),
                "action": AuditActionType.UPDATED.value,
                "user_id": user_id,
                "user_name": user_name,
                "changes": changes,
                "created_at": datetime.utcnow()
            })
        
        logger.info(f"Booking edited: {booking_id} by {user_name or 'system'}")
        
        # Fetch updated booking
        booking_doc = db.bookings.find_one({"_id": ObjectId(booking_id)})
        return BookingService._format_booking_response(booking_doc)
    
    @staticmethod
    def reschedule_booking(
        booking_id: str,
        tenant_id: str,
        new_date: str,
        new_stylist_id: Optional[str] = None,
        user_id: Optional[str] = None,
        user_name: Optional[str] = None
    ) -> Dict:
        """
        Reschedule a booking to a new date/time
        
        Requirements: 2.3, 2.4
        
        Args:
            booking_id: Booking ID to reschedule
            tenant_id: Tenant ID
            new_date: New booking date/time in ISO format
            new_stylist_id: Optional new stylist ID
            user_id: User performing the reschedule
            user_name: Name of user performing the reschedule
            
        Returns:
            Dict with rescheduled booking data
        """
        db = Database.get_db()
        
        # Get existing booking
        booking_doc = db.bookings.find_one({
            "_id": ObjectId(booking_id),
            "tenant_id": ObjectId(tenant_id)
        })
        
        if not booking_doc:
            raise NotFoundException("Booking not found")
        
        # Parse new date
        try:
            new_datetime = datetime.fromisoformat(new_date.replace('Z', '+00:00'))
        except ValueError:
            try:
                new_datetime = datetime.strptime(new_date, '%Y-%m-%d')
            except ValueError:
                raise BadRequestException("Invalid date format")
        
        # Prepare update data
        update_data = {
            "booking_date": new_datetime,
            "updated_at": datetime.utcnow()
        }
        
        # Track changes for audit log
        changes = {
            "booking_date": {
                "old": booking_doc["booking_date"].isoformat() if isinstance(booking_doc["booking_date"], datetime) else str(booking_doc["booking_date"]),
                "new": new_datetime.isoformat()
            }
        }
        
        # Handle stylist change if provided
        if new_stylist_id:
            stylist = db.stylists.find_one({
                "_id": ObjectId(new_stylist_id),
                "tenant_id": ObjectId(tenant_id)
            })
            if not stylist:
                raise NotFoundException("Stylist not found")
            
            changes["stylist_id"] = {
                "old": str(booking_doc["stylist_id"]),
                "new": new_stylist_id
            }
            update_data["stylist_id"] = ObjectId(new_stylist_id)
            update_data["stylist_name"] = stylist.get("name")
        
        # TODO: Check for conflicts at new date/time
        # This will be implemented in Task 1.5
        
        # Update booking
        db.bookings.update_one(
            {"_id": ObjectId(booking_id)},
            {"$set": update_data}
        )
        
        # Create audit log
        from app.schemas.booking_audit import AuditActionType
        db.booking_audit_logs.insert_one({
            "booking_id": booking_id,
            "tenant_id": ObjectId(tenant_id),
            "action": AuditActionType.RESCHEDULED.value,
            "user_id": user_id,
            "user_name": user_name,
            "changes": changes,
            "created_at": datetime.utcnow()
        })
        
        logger.info(f"Booking rescheduled: {booking_id} to {new_datetime} by {user_name or 'system'}")
        
        # TODO: Send notification to client about reschedule
        # This will be implemented in Task 2.5
        
        # Fetch updated booking
        booking_doc = db.bookings.find_one({"_id": ObjectId(booking_id)})
        return BookingService._format_booking_response(booking_doc)
    
    @staticmethod
    def _format_booking_response(booking_doc: Dict) -> Dict:
        """Format booking document for response"""
        db = Database.get_db()
        
        # Get location name if location_id exists
        location_name = "Unknown Location"
        if booking_doc.get("location_id"):
            location = db.locations.find_one({
                "_id": ObjectId(booking_doc["location_id"]) if isinstance(booking_doc["location_id"], str) else booking_doc["location_id"]
            })
            if location:
                location_name = location.get("name", "Unknown Location")
        
        return {
            "id": str(booking_doc["_id"]),
            "tenant_id": str(booking_doc["tenant_id"]) if isinstance(booking_doc["tenant_id"], ObjectId) else booking_doc["tenant_id"],
            "location_id": str(booking_doc.get("location_id", "")) if booking_doc.get("location_id") else "",
            "location_name": location_name,
            "client_id": str(booking_doc["client_id"]) if isinstance(booking_doc.get("client_id"), ObjectId) else booking_doc.get("client_id"),
            "client_name": booking_doc.get("client_name", "Unknown"),
            "client_phone": booking_doc.get("client_phone", ""),
            "service_id": str(booking_doc["service_id"]) if isinstance(booking_doc["service_id"], ObjectId) else booking_doc["service_id"],
            "service_name": booking_doc.get("service_name", "Unknown Service"),
            "service_price": booking_doc.get("service_price", 0.0),
            "stylist_id": str(booking_doc["stylist_id"]) if isinstance(booking_doc["stylist_id"], ObjectId) else booking_doc["stylist_id"],
            "stylist_name": booking_doc.get("stylist_name", "Unknown Stylist"),
            "booking_date": booking_doc["booking_date"],
            "duration_minutes": booking_doc.get("duration_minutes", 30),
            "status": booking_doc.get("status", "pending"),
            "payment_status": booking_doc.get("payment_status", "unpaid"),
            "payment_method": booking_doc.get("payment_method"),
            "paid_via_package_credit": booking_doc.get("paid_via_package_credit", False),
            "package_credit_id": booking_doc.get("package_credit_id"),
            "notes": booking_doc.get("notes"),
            "created_at": booking_doc.get("created_at", datetime.utcnow()),
            "updated_at": booking_doc.get("updated_at", datetime.utcnow()),
            "confirmed_at": booking_doc.get("confirmed_at"),
            "completed_at": booking_doc.get("completed_at"),
            "cancelled_at": booking_doc.get("cancelled_at")
        }
    
    @staticmethod
    async def export_bookings_csv(
        tenant_id: str,
        status: Optional[str] = None,
        stylist_id: Optional[str] = None,
        client_id: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        booking_type: Optional[str] = None,
        group_subtype: Optional[str] = None
    ) -> str:
        """
        Export bookings to CSV format
        
        Returns:
            CSV string
        """
        import csv
        import io
        
        # Get all bookings with filters (no pagination)
        bookings = await BookingService.get_bookings(
            tenant_id=tenant_id,
            status=status,
            stylist_id=stylist_id,
            client_id=client_id,
            date_from=date_from,
            date_to=date_to,
            booking_type=booking_type,
            group_subtype=group_subtype,
            offset=0,
            limit=10000  # Large limit for export
        )
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'Booking ID', 'Client Name', 'Client Phone', 'Service', 'Stylist',
            'Date', 'Time', 'Duration (mins)', 'Status', 'Payment Status',
            'Payment Method', 'Price', 'Notes', 'Created At'
        ])
        
        # Write data
        for booking in bookings:
            booking_date = booking['booking_date']
            if isinstance(booking_date, str):
                booking_date = datetime.fromisoformat(booking_date.replace('Z', '+00:00'))
            
            writer.writerow([
                booking['id'],
                booking['client_name'],
                booking['client_phone'],
                booking['service_name'],
                booking['stylist_name'],
                booking_date.strftime('%Y-%m-%d'),
                booking_date.strftime('%H:%M'),
                booking['duration_minutes'],
                booking['status'],
                booking.get('payment_status', 'unpaid'),
                booking.get('payment_method', ''),
                booking.get('service_price', 0),
                booking.get('notes', ''),
                booking['created_at'].strftime('%Y-%m-%d %H:%M') if booking.get('created_at') else ''
            ])
        
        return output.getvalue()
    
    @staticmethod
    async def export_bookings_excel(
        tenant_id: str,
        status: Optional[str] = None,
        stylist_id: Optional[str] = None,
        client_id: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        booking_type: Optional[str] = None,
        group_subtype: Optional[str] = None
    ) -> bytes:
        """
        Export bookings to Excel format
        
        Returns:
            Excel file bytes
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        import io
        
        # Get all bookings with filters
        bookings = await BookingService.get_bookings(
            tenant_id=tenant_id,
            status=status,
            stylist_id=stylist_id,
            client_id=client_id,
            date_from=date_from,
            date_to=date_to,
            booking_type=booking_type,
            group_subtype=group_subtype,
            offset=0,
            limit=10000
        )
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Write header
        headers = [
            'Booking ID', 'Client Name', 'Client Phone', 'Service', 'Stylist',
            'Date', 'Time', 'Duration (mins)', 'Status', 'Payment Status',
            'Payment Method', 'Price', 'Notes', 'Created At'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data
        for row, booking in enumerate(bookings, 2):
            booking_date = booking['booking_date']
            if isinstance(booking_date, str):
                booking_date = datetime.fromisoformat(booking_date.replace('Z', '+00:00'))
            
            ws.cell(row=row, column=1, value=booking['id'])
            ws.cell(row=row, column=2, value=booking['client_name'])
            ws.cell(row=row, column=3, value=booking['client_phone'])
            ws.cell(row=row, column=4, value=booking['service_name'])
            ws.cell(row=row, column=5, value=booking['stylist_name'])
            ws.cell(row=row, column=6, value=booking_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=7, value=booking_date.strftime('%H:%M'))
            ws.cell(row=row, column=8, value=booking['duration_minutes'])
            ws.cell(row=row, column=9, value=booking['status'])
            ws.cell(row=row, column=10, value=booking.get('payment_status', 'unpaid'))
            ws.cell(row=row, column=11, value=booking.get('payment_method', ''))
            ws.cell(row=row, column=12, value=booking.get('service_price', 0))
            ws.cell(row=row, column=13, value=booking.get('notes', ''))
            ws.cell(row=row, column=14, value=booking['created_at'].strftime('%Y-%m-%d %H:%M') if booking.get('created_at') else '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


    @staticmethod
    async def check_available_package_credits(
        client_id: str,
        service_id: str,
        tenant_id: str
    ) -> List[Dict]:
        """
        Check available package credits for a client and service
        
        Requirements: 14.1, 14.2
        
        Returns:
            List of available package credits
        """
        db = Database.get_db()
        
        try:
            # Get active packages for client with available credits for this service
            packages = list(db.package_purchases.find({
                "client_id": client_id,
                "tenant_id": tenant_id,
                "status": "active"
            }))
            
            available_credits = []
            
            for package in packages:
                # Check if package is expired
                if package.get("expiration_date"):
                    if datetime.utcnow() > package["expiration_date"]:
                        continue
                
                # Get service credits for this service
                credits = list(db.service_credits.find({
                    "purchase_id": str(package["_id"]),
                    "service_id": service_id,
                    "remaining_quantity": {"$gt": 0},
                    "status": {"$in": ["available", "reserved"]}
                }))
                
                for credit in credits:
                    available_credits.append({
                        "credit_id": str(credit["_id"]),
                        "purchase_id": str(package["_id"]),
                        "package_name": db.packages.find_one({"_id": ObjectId(package["package_definition_id"])}).get("name", "Unknown Package") if package.get("package_definition_id") else "Unknown Package",
                        "service_id": service_id,
                        "remaining_quantity": credit.get("remaining_quantity", 0),
                        "service_price": credit.get("service_price", 0),
                        "expiration_date": package.get("expiration_date")
                    })
            
            return available_credits
        
        except Exception as e:
            logger.error(f"Error checking available package credits: {e}")
            raise Exception(f"Failed to check available package credits: {str(e)}")
    
    @staticmethod
    async def reserve_package_credit_for_booking(
        booking_id: str,
        credit_id: str,
        tenant_id: str
    ) -> Dict:
        """
        Reserve a package credit for a booking
        
        Requirements: 14.3
        
        Returns:
            Reservation details
        """
        db = Database.get_db()
        
        try:
            # Get booking
            booking = db.bookings.find_one({
                "_id": ObjectId(booking_id),
                "tenant_id": tenant_id
            })
            
            if not booking:
                raise ValueError("Booking not found")
            
            # Get credit
            credit = db.service_credits.find_one({
                "_id": ObjectId(credit_id)
            })
            
            if not credit:
                raise ValueError("Credit not found")
            
            # Check if credit has remaining quantity
            if credit.get("remaining_quantity", 0) < 1:
                raise ValueError("No remaining credits available")
            
            # Create reservation
            reservation_data = {
                "credit_id": credit_id,
                "booking_id": booking_id,
                "reserved_at": datetime.utcnow(),
                "expires_at": datetime.utcnow() + timedelta(hours=24),
                "status": "active"
            }
            
            result = db.credit_reservations.insert_one(reservation_data)
            reservation_data["_id"] = str(result.inserted_id)
            
            # Update credit reserved quantity
            db.service_credits.update_one(
                {"_id": ObjectId(credit_id)},
                {
                    "$inc": {"reserved_quantity": 1},
                    "$set": {"updated_at": datetime.utcnow()}
                }
            )
            
            # Update booking with package credit info
            db.bookings.update_one(
                {"_id": ObjectId(booking_id)},
                {
                    "$set": {
                        "package_credit_id": credit_id,
                        "package_credit_reservation_id": str(result.inserted_id),
                        "paid_via_package_credit": True,
                        "payment_status": "paid",
                        "updated_at": datetime.utcnow()
                    }
                }
            )
            
            logger.info(f"Package credit reserved for booking: {booking_id}")
            
            return reservation_data
        
        except ValueError as e:
            raise e
        except Exception as e:
            logger.error(f"Error reserving package credit for booking: {e}")
            raise Exception(f"Failed to reserve package credit: {str(e)}")
    
    @staticmethod
    async def redeem_package_credit_for_booking(
        booking_id: str,
        tenant_id: str,
        staff_id: str
    ) -> Dict:
        """
        Redeem a reserved package credit when booking is completed
        
        Requirements: 14.4
        
        Returns:
            Redemption details
        """
        db = Database.get_db()
        
        try:
            # Get booking
            booking = db.bookings.find_one({
                "_id": ObjectId(booking_id),
                "tenant_id": tenant_id
            })
            
            if not booking:
                raise ValueError("Booking not found")
            
            # Check if booking has reserved package credit
            if not booking.get("package_credit_id"):
                raise ValueError("Booking does not have reserved package credit")
            
            # Get credit
            credit = db.service_credits.find_one({
                "_id": ObjectId(booking["package_credit_id"])
            })
            
            if not credit:
                raise ValueError("Credit not found")
            
            # Create redemption transaction
            redemption_data = {
                "purchase_id": credit["purchase_id"],
                "credit_id": booking["package_credit_id"],
                "service_id": credit["service_id"],
                "client_id": booking["client_id"],
                "redeemed_by_staff_id": staff_id,
                "redemption_date": datetime.utcnow(),
                "service_value": credit.get("service_price", 0),
                "booking_id": booking_id,
                "created_at": datetime.utcnow()
            }
            
            result = db.redemption_transactions.insert_one(redemption_data)
            redemption_data["_id"] = str(result.inserted_id)
            
            # Update credit remaining quantity
            new_remaining = credit.get("remaining_quantity", 0) - 1
            new_status = "redeemed" if new_remaining == 0 else credit.get("status", "available")
            
            db.service_credits.update_one(
                {"_id": ObjectId(booking["package_credit_id"])},
                {
                    "$inc": {"remaining_quantity": -1, "reserved_quantity": -1},
                    "$set": {
                        "status": new_status,
                        "updated_at": datetime.utcnow()
                    }
                }
            )
            
            # Check if all credits in package are redeemed
            remaining_credits = db.service_credits.find_one({
                "purchase_id": credit["purchase_id"],
                "remaining_quantity": {"$gt": 0}
            })
            
            if not remaining_credits:
                # Mark package as fully redeemed
                db.package_purchases.update_one(
                    {"_id": ObjectId(credit["purchase_id"])},
                    {
                        "$set": {
                            "status": "fully_redeemed",
                            "updated_at": datetime.utcnow()
                        }
                    }
                )
            
            logger.info(f"Package credit redeemed for booking: {booking_id}")
            
            return redemption_data
        
        except ValueError as e:
            raise e
        except Exception as e:
            logger.error(f"Error redeeming package credit for booking: {e}")
            raise Exception(f"Failed to redeem package credit: {str(e)}")
    
    @staticmethod
    async def release_package_credit_for_booking(
        booking_id: str,
        tenant_id: str
    ) -> Dict:
        """
        Release a reserved package credit when booking is cancelled
        
        Requirements: 14.5
        
        Returns:
            Release details
        """
        db = Database.get_db()
        
        try:
            # Get booking
            booking = db.bookings.find_one({
                "_id": ObjectId(booking_id),
                "tenant_id": tenant_id
            })
            
            if not booking:
                raise ValueError("Booking not found")
            
            # Check if booking has reserved package credit
            if not booking.get("package_credit_id"):
                raise ValueError("Booking does not have reserved package credit")
            
            # Get credit
            credit = db.service_credits.find_one({
                "_id": ObjectId(booking["package_credit_id"])
            })
            
            if not credit:
                raise ValueError("Credit not found")
            
            # Get reservation
            reservation = db.credit_reservations.find_one({
                "_id": ObjectId(booking.get("package_credit_reservation_id"))
            })
            
            if reservation:
                # Update reservation status
                db.credit_reservations.update_one(
                    {"_id": ObjectId(booking["package_credit_reservation_id"])},
                    {
                        "$set": {"status": "released"}
                    }
                )
            
            # Update credit reserved quantity
            db.service_credits.update_one(
                {"_id": ObjectId(booking["package_credit_id"])},
                {
                    "$inc": {"reserved_quantity": -1},
                    "$set": {
                        "status": "available",
                        "updated_at": datetime.utcnow()
                    }
                }
            )
            
            # Update booking to remove package credit info
            db.bookings.update_one(
                {"_id": ObjectId(booking_id)},
                {
                    "$unset": {
                        "package_credit_id": "",
                        "package_credit_reservation_id": "",
                        "paid_via_package_credit": ""
                    },
                    "$set": {
                        "payment_status": "unpaid",
                        "updated_at": datetime.utcnow()
                    }
                }
            )
            
            logger.info(f"Package credit released for booking: {booking_id}")
            
            return {
                "booking_id": booking_id,
                "credit_id": booking["package_credit_id"],
                "status": "released"
            }
        
        except ValueError as e:
            raise e
        except Exception as e:
            logger.error(f"Error releasing package credit for booking: {e}")
            raise Exception(f"Failed to release package credit: {str(e)}")
    
    @staticmethod
    def bulk_actions(
        tenant_id: str,
        action: str,
        booking_ids: List[str],
        data: Optional[Dict] = None
    ) -> Dict:
        """
        Perform bulk actions on multiple bookings
        
        Args:
            tenant_id: Tenant ID
            action: Action to perform (confirm, cancel, complete, etc.)
            booking_ids: List of booking IDs
            data: Additional data for the action
            
        Returns:
            Dict with results
        """
        db = Database.get_db()
        
        results = {
            "action": action,
            "total": len(booking_ids),
            "successful": 0,
            "failed": 0,
            "errors": []
        }
        
        for booking_id in booking_ids:
            try:
                if action == "confirm":
                    BookingService.update_booking_status(booking_id, tenant_id, "confirmed")
                    results["successful"] += 1
                elif action == "cancel":
                    BookingService.update_booking_status(booking_id, tenant_id, "cancelled")
                    results["successful"] += 1
                elif action == "complete":
                    BookingService.update_booking_status(booking_id, tenant_id, "completed")
                    results["successful"] += 1
                else:
                    results["failed"] += 1
                    results["errors"].append(f"Unknown action: {action}")
            except Exception as e:
                results["failed"] += 1
                results["errors"].append(f"Booking {booking_id}: {str(e)}")
        
        logger.info(f"Bulk action '{action}' completed: {results['successful']} successful, {results['failed']} failed")
        return results
    
    @staticmethod
    def create_recurring_booking(
        tenant_id: str,
        data: Dict
    ) -> Dict:
        """
        Create a recurring booking template
        
        Args:
            tenant_id: Tenant ID
            data: Booking template data
            
        Returns:
            Dict with created template
        """
        db = Database.get_db()
        
        template_data = {
            "tenant_id": tenant_id,
            "client_id": data.get("client_id"),
            "stylist_id": data.get("stylist_id"),
            "service_id": data.get("service_id"),
            "start_date": datetime.fromisoformat(data.get("start_date", "").replace('Z', '+00:00')),
            "frequency": data.get("frequency", "weekly"),  # daily, weekly, monthly
            "end_date": datetime.fromisoformat(data.get("end_date", "").replace('Z', '+00:00')) if data.get("end_date") else None,
            "occurrence_count": data.get("occurrence_count", 4),
            "notes": data.get("notes"),
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }
        
        result = db.booking_templates.insert_one(template_data)
        template_data["_id"] = str(result.inserted_id)
        
        logger.info(f"Recurring booking template created: {result.inserted_id}")
        return template_data
    
    @staticmethod
    def update_recurring_booking(
        tenant_id: str,
        template_id: str,
        update_type: str,
        booking_id: Optional[str] = None,
        data: Optional[Dict] = None
    ) -> Dict:
        """
        Update a recurring booking template
        
        Args:
            tenant_id: Tenant ID
            template_id: Template ID
            update_type: single, future, or all
            booking_id: Specific booking ID if update_type is single
            data: Update data
            
        Returns:
            Dict with results
        """
        db = Database.get_db()
        
        template = db.booking_templates.find_one({
            "_id": ObjectId(template_id),
            "tenant_id": tenant_id
        })
        
        if not template:
            raise NotFoundException("Booking template not found")
        
        if update_type == "single" and booking_id:
            # Update single booking
            db.bookings.update_one(
                {"_id": ObjectId(booking_id)},
                {"$set": {**data, "updated_at": datetime.utcnow()}}
            )
        elif update_type == "future":
            # Update this and future bookings
            booking = db.bookings.find_one({"_id": ObjectId(booking_id)})
            if booking:
                db.bookings.update_many(
                    {
                        "booking_template_id": template_id,
                        "booking_date": {"$gte": booking["booking_date"]}
                    },
                    {"$set": {**data, "updated_at": datetime.utcnow()}}
                )
        elif update_type == "all":
            # Update all bookings in series
            db.bookings.update_many(
                {"booking_template_id": template_id},
                {"$set": {**data, "updated_at": datetime.utcnow()}}
            )
        
        logger.info(f"Recurring booking updated: {template_id} ({update_type})")
        return {"template_id": template_id, "update_type": update_type, "status": "updated"}
    
    @staticmethod
    def delete_recurring_booking(
        tenant_id: str,
        template_id: str,
        update_type: str,
        booking_id: Optional[str] = None
    ) -> Dict:
        """
        Delete a recurring booking template
        
        Args:
            tenant_id: Tenant ID
            template_id: Template ID
            update_type: single, future, or all
            booking_id: Specific booking ID if update_type is single
            
        Returns:
            Dict with results
        """
        db = Database.get_db()
        
        template = db.booking_templates.find_one({
            "_id": ObjectId(template_id),
            "tenant_id": tenant_id
        })
        
        if not template:
            raise NotFoundException("Booking template not found")
        
        if update_type == "single" and booking_id:
            # Delete single booking
            db.bookings.delete_one({"_id": ObjectId(booking_id)})
        elif update_type == "future":
            # Delete this and future bookings
            booking = db.bookings.find_one({"_id": ObjectId(booking_id)})
            if booking:
                db.bookings.delete_many({
                    "booking_template_id": template_id,
                    "booking_date": {"$gte": booking["booking_date"]}
                })
        elif update_type == "all":
            # Delete all bookings in series and template
            db.bookings.delete_many({"booking_template_id": template_id})
            db.booking_templates.delete_one({"_id": ObjectId(template_id)})
        
        logger.info(f"Recurring booking deleted: {template_id} ({update_type})")
        return {"template_id": template_id, "update_type": update_type, "status": "deleted"}
    
    @staticmethod
    def get_booking_analytics(
        tenant_id: str,
        date_from: str,
        date_to: str,
        group_subtype: Optional[str] = None
    ) -> Dict:
        """
        Get booking analytics for a date range
        
        Args:
            tenant_id: Tenant ID
            date_from: Start date (YYYY-MM-DD)
            date_to: End date (YYYY-MM-DD)
            group_subtype: Optional group subtype filter
            
        Returns:
            Dict with analytics data
        """
        db = Database.get_db()
        
        # Parse dates
        try:
            date_from_dt = datetime.strptime(date_from, '%Y-%m-%d')
            date_to_dt = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
        except ValueError:
            raise BadRequestException("Invalid date format. Use YYYY-MM-DD")
        
        # Build query
        query = {
            "tenant_id": tenant_id,
            "booking_date": {
                "$gte": date_from_dt,
                "$lte": date_to_dt
            }
        }
        
        if group_subtype:
            query["group_subtype"] = group_subtype
        
        # Get bookings
        bookings = list(db.bookings.find(query))
        
        # Calculate analytics
        total_bookings = len(bookings)
        completed = len([b for b in bookings if b.get("status") == "completed"])
        cancelled = len([b for b in bookings if b.get("status") == "cancelled"])
        pending = len([b for b in bookings if b.get("status") == "pending"])
        confirmed = len([b for b in bookings if b.get("status") == "confirmed"])
        
        total_revenue = sum(b.get("service_price", 0) for b in bookings if b.get("status") == "completed")
        
        return {
            "date_range": {
                "from": date_from,
                "to": date_to
            },
            "total_bookings": total_bookings,
            "by_status": {
                "completed": completed,
                "cancelled": cancelled,
                "pending": pending,
                "confirmed": confirmed
            },
            "completion_rate": (completed / total_bookings * 100) if total_bookings > 0 else 0,
            "total_revenue": total_revenue,
            "average_booking_value": (total_revenue / completed) if completed > 0 else 0
        }
    
    @staticmethod
    def get_booking_audit_log(
        booking_id: str,
        tenant_id: str
    ) -> Dict:
        """
        Get audit log for a booking
        
        Args:
            booking_id: Booking ID
            tenant_id: Tenant ID
            
        Returns:
            Dict with audit log entries
        """
        db = Database.get_db()
        
        # Verify booking exists
        booking = db.bookings.find_one({
            "_id": ObjectId(booking_id),
            "tenant_id": tenant_id
        })
        
        if not booking:
            raise NotFoundException("Booking not found")
        
        # Get audit logs
        logs = list(db.booking_audit_logs.find({
            "booking_id": booking_id,
            "tenant_id": tenant_id
        }).sort("created_at", -1))
        
        # Convert ObjectIds to strings
        for log in logs:
            log["_id"] = str(log["_id"])
            log["tenant_id"] = str(log["tenant_id"])
        
        return {
            "booking_id": booking_id,
            "items": logs,
            "total": len(logs)
        }


# Singleton instance
booking_service = BookingService()
